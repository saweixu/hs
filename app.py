import io
import re
import json
from datetime import date
from collections import defaultdict

import pandas as pd
import requests
import streamlit as st
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

# =========================================================
# CONFIG
# =========================================================
DEFAULT_COUNTRY = "CN"
TODAY_ISO = date.today().isoformat()

st.set_page_config(page_title="HS Code Analyzer", layout="wide")
st.title("HS Code Analyzer from Invoices")
st.caption(
    "Upload multiple invoice files, extract HS codes from INVOICE column C, "
    "remove duplicates, analyze them through a JSON API, and export OUTPUT.xlsx"
)

# =========================================================
# SIDEBAR - API CONFIG
# =========================================================
with st.sidebar:
    st.header("API Settings")

    default_base_url = st.secrets.get("TARIC_SUPPORT_BASE_URL", "https://api.taricsupport.com")
    default_token = st.secrets.get("TARIC_SUPPORT_TOKEN", "")
    default_endpoint = st.secrets.get("TARIC_SUPPORT_MEASURES_ENDPOINT", "")

    base_url = st.text_input("Base URL", value=default_base_url)
    api_token = st.text_input("API Token", value=default_token, type="password")
    measures_endpoint = st.text_input(
        "Measures endpoint path",
        value=default_endpoint,
        help="Example: /v2/xxxxx  ← copy the exact path from Taric Support Swagger"
    )

    st.markdown(
        "Use the exact endpoint path shown in Taric Support Swagger for the tariff/measures lookup."
    )

# =========================================================
# HELPERS - EXCEL
# =========================================================
def find_sum_row(ws, search_col=2):
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, search_col).value
        if v is None:
            continue
        txt = str(v).strip().upper().replace(":", "").replace(" ", "")
        if txt == "SUM":
            return r
    return None


def get_merged_cell_value(ws, row, col):
    cell = ws.cell(row=row, column=col)

    if not isinstance(cell, MergedCell):
        return cell.value

    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return ws.cell(merged_range.min_row, merged_range.min_col).value

    return None


def get_best_cell_value(ws_data, ws_formula, row, col):
    v_data = get_merged_cell_value(ws_data, row, col)
    if v_data not in (None, ""):
        return v_data
    return get_merged_cell_value(ws_formula, row, col)


def normalize_hs_code(raw):
    if raw is None:
        return []

    text = str(raw).strip()
    candidates = set()

    # direct 6-10 digit blocks
    for m in re.findall(r"(?<!\d)(\d{6,10})(?!\d)", text):
        candidates.add(m)

    # dotted / spaced / slashed groups
    for m in re.findall(r"(?<!\d)(\d{4}[.\s/-]?\d{2}(?:[.\s/-]?\d{2}){0,2})(?!\d)", text):
        digits = re.sub(r"\D", "", m)
        if 6 <= len(digits) <= 10:
            candidates.add(digits)

    return sorted(candidates)


def extract_hs_from_invoice_file(uploaded_file):
    uploaded_file.seek(0)
    file_bytes = uploaded_file.read()

    wb_data = load_workbook(io.BytesIO(file_bytes), data_only=True)
    wb_formula = load_workbook(io.BytesIO(file_bytes), data_only=False)

    if "INVOICE" in wb_data.sheetnames:
        ws_data = wb_data["INVOICE"]
        ws_formula = wb_formula["INVOICE"]
        sheet_used = "INVOICE"
    else:
        sheet_used = wb_data.sheetnames[0]
        ws_data = wb_data[sheet_used]
        ws_formula = wb_formula[sheet_used]

    sum_row = find_sum_row(ws_data, search_col=2)
    if not sum_row:
        sum_row = find_sum_row(ws_formula, search_col=2)

    if not sum_row:
        return [], [], f"{uploaded_file.name}: SUM row not found in sheet '{sheet_used}'"

    results = []
    debug_rows = []

    for row in range(20, sum_row):
        cell_value = get_best_cell_value(ws_data, ws_formula, row, 3)  # column C

        debug_rows.append({
            "file_name": uploaded_file.name,
            "sheet_name": sheet_used,
            "row": row,
            "cell": f"C{row}",
            "raw_value": "" if cell_value is None else str(cell_value),
        })

        codes = normalize_hs_code(cell_value)
        for code in codes:
            results.append({
                "file_name": uploaded_file.name,
                "sheet_name": sheet_used,
                "row": row,
                "raw_cell_value": "" if cell_value is None else str(cell_value),
                "hs_code": code,
            })

    return results, debug_rows, None


# =========================================================
# HELPERS - API JSON
# =========================================================
def build_full_url(base_url: str, endpoint_path: str) -> str:
    return base_url.rstrip("/") + "/" + endpoint_path.lstrip("/")


def get_api_headers(token: str) -> dict:
    return {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
        "Content-Type": "application/json",
        "User-Agent": "Mozilla/5.0",
    }


def safe_json(value):
    try:
        return json.dumps(value, ensure_ascii=False, indent=2)
    except Exception:
        return str(value)


def flatten_strings(obj, found=None):
    if found is None:
        found = []

    if obj is None:
        return found

    if isinstance(obj, dict):
        for v in obj.values():
            flatten_strings(v, found)
    elif isinstance(obj, list):
        for v in obj:
            flatten_strings(v, found)
    else:
        found.append(str(obj))

    return found


def shorten_text(text, max_len=3000):
    if text is None:
        return ""
    text = str(text).strip()
    if len(text) <= max_len:
        return text
    return text[:max_len] + " ..."


def extract_description_from_json(data):
    """
    Best-effort extraction.
    """
    candidate_keys = [
        "description", "goodsDescription", "goods_description",
        "productDescription", "tariffDescription", "descriptionEn"
    ]

    if isinstance(data, dict):
        for key in candidate_keys:
            if key in data and data[key]:
                return str(data[key])

        for value in data.values():
            desc = extract_description_from_json(value)
            if desc:
                return desc

    elif isinstance(data, list):
        for item in data:
            desc = extract_description_from_json(item)
            if desc:
                return desc

    return ""


def extract_duty_from_json(data):
    """
    Best-effort extraction of duty/rate related values.
    """
    lines = []
    seen = set()
    keywords = (
        "duty", "rate", "third country", "erga omnes",
        "customs", "%", "ad valorem"
    )

    for item in flatten_strings(data, []):
        text = " ".join(item.split())
        low = text.lower()
        if any(k in low for k in keywords):
            if text not in seen:
                seen.add(text)
                lines.append(text)

    return " | ".join(lines[:15])


def extract_restrictions_from_json(data):
    lines = []
    seen = set()
    keywords = (
        "restriction", "prohibition", "certificate", "license", "licence",
        "document", "quota", "anti-dumping", "measure", "additional code"
    )

    for item in flatten_strings(data, []):
        text = " ".join(item.split())
        low = text.lower()
        if any(k in low for k in keywords):
            if text not in seen:
                seen.add(text)
                lines.append(text)

    return " | ".join(lines[:20])


def analyze_hs_code_via_json_api(
    hs_code: str,
    country_code: str,
    reference_date: str,
    base_url: str,
    endpoint_path: str,
    token: str,
):
    """
    Generic JSON POST request.
    You may need to adjust the payload keys to match the exact Taric Support endpoint.
    """
    if not token.strip():
        return {
            "hs_code": hs_code,
            "country_code": country_code,
            "reference_date": reference_date,
            "description_en": "",
            "duty_rate": "",
            "restrictions_documents": "",
            "status": "ERROR",
            "error": "Missing API token.",
            "raw_response": "",
        }

    if not endpoint_path.strip():
        return {
            "hs_code": hs_code,
            "country_code": country_code,
            "reference_date": reference_date,
            "description_en": "",
            "duty_rate": "",
            "restrictions_documents": "",
            "status": "ERROR",
            "error": "Missing endpoint path. Paste the exact measures endpoint from Swagger.",
            "raw_response": "",
        }

    url = build_full_url(base_url, endpoint_path)
    headers = get_api_headers(token)

    # Payload générique.
    # Si la doc réelle utilise d'autres noms de champs, il faudra juste adapter ici.
    payload = {
        "goodsCode": hs_code,
        "countryCode": country_code,
        "referenceDate": reference_date,
        "tradeMovement": "IMPORT",
        "languageCode": "EN",
    }

    try:
        response = requests.post(url, headers=headers, json=payload, timeout=45)
        raw_text = response.text

        if response.status_code >= 400:
            return {
                "hs_code": hs_code,
                "country_code": country_code,
                "reference_date": reference_date,
                "description_en": "",
                "duty_rate": "",
                "restrictions_documents": "",
                "status": "ERROR",
                "error": f"HTTP {response.status_code}: {raw_text[:500]}",
                "raw_response": raw_text[:12000],
            }

        try:
            data = response.json()
        except Exception:
            return {
                "hs_code": hs_code,
                "country_code": country_code,
                "reference_date": reference_date,
                "description_en": "",
                "duty_rate": "",
                "restrictions_documents": "",
                "status": "ERROR",
                "error": "API did not return valid JSON.",
                "raw_response": raw_text[:12000],
            }

        description = shorten_text(extract_description_from_json(data), 1000)
        duty_rate = shorten_text(extract_duty_from_json(data), 2000)
        restrictions = shorten_text(extract_restrictions_from_json(data), 3000)

        return {
            "hs_code": hs_code,
            "country_code": country_code,
            "reference_date": reference_date,
            "description_en": description,
            "duty_rate": duty_rate,
            "restrictions_documents": restrictions,
            "status": "OK",
            "error": "",
            "raw_response": safe_json(data)[:12000],
        }

    except requests.RequestException as e:
        return {
            "hs_code": hs_code,
            "country_code": country_code,
            "reference_date": reference_date,
            "description_en": "",
            "duty_rate": "",
            "restrictions_documents": "",
            "status": "ERROR",
            "error": str(e),
            "raw_response": "",
        }


# =========================================================
# OUTPUT EXCEL
# =========================================================
def build_output_excel(df_hs_found, df_summary, df_debug):
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_hs_found.to_excel(writer, index=False, sheet_name="HS_FOUND")
        df_summary.to_excel(writer, index=False, sheet_name="OUTPUT")
        df_debug.to_excel(writer, index=False, sheet_name="DEBUG_READ")

    output.seek(0)
    return output


# =========================================================
# UI
# =========================================================
uploaded_files = st.file_uploader(
    "Upload invoice files (.xlsx)",
    type=["xlsx"],
    accept_multiple_files=True
)

if uploaded_files:
    st.info(f"{len(uploaded_files)} file(s) uploaded.")

    all_rows = []
    all_debug_rows = []
    warnings = []

    for f in uploaded_files:
        rows, debug_rows, err = extract_hs_from_invoice_file(f)
        if err:
            warnings.append(err)
        all_rows.extend(rows)
        all_debug_rows.extend(debug_rows)

    if warnings:
        for w in warnings:
            st.warning(w)

    if not all_rows:
        st.error("No HS code found in uploaded files.")
        if all_debug_rows:
            with st.expander("Debug preview", expanded=True):
                st.dataframe(pd.DataFrame(all_debug_rows), use_container_width=True)
        st.stop()

    df_found = pd.DataFrame(all_rows)
    df_debug = pd.DataFrame(all_debug_rows)

    grouped_files = defaultdict(set)
    grouped_positions = defaultdict(list)

    for row in all_rows:
        grouped_files[row["hs_code"]].add(row["file_name"])
        grouped_positions[row["hs_code"]].append(f"{row['file_name']} [row {row['row']}]")

    unique_codes = sorted(df_found["hs_code"].dropna().astype(str).unique().tolist())

    st.success(f"{len(df_found)} HS code occurrence(s) found.")
    st.success(f"{len(unique_codes)} unique HS code(s) after duplicate removal.")

    with st.expander("Preview extracted HS codes", expanded=False):
        st.dataframe(df_found, use_container_width=True)

    with st.expander("Debug preview (read values from column C)", expanded=False):
        st.dataframe(df_debug, use_container_width=True)

    if st.button("Analyze HS Codes"):
        summary_rows = []
        progress = st.progress(0)
        total = len(unique_codes)

        for i, hs in enumerate(unique_codes, start=1):
            result = analyze_hs_code_via_json_api(
                hs_code=hs,
                country_code=DEFAULT_COUNTRY,
                reference_date=TODAY_ISO,
                base_url=base_url,
                endpoint_path=measures_endpoint,
                token=api_token,
            )

            result["source_file_count"] = len(grouped_files[hs])
            result["source_files"] = " | ".join(sorted(grouped_files[hs]))
            result["source_positions"] = " | ".join(grouped_positions[hs])

            summary_rows.append(result)
            progress.progress(i / total)

        df_summary = pd.DataFrame(summary_rows)

        st.subheader("Analysis Result")
        st.dataframe(df_summary, use_container_width=True)

        xlsx_data = build_output_excel(df_found, df_summary, df_debug)

        st.download_button(
            label="Download OUTPUT.xlsx",
            data=xlsx_data,
            file_name="OUTPUT.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

else:
    st.info("Upload one or more invoice files to begin.")
