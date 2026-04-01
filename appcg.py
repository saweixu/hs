import io
import re
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# =========================================================
# CONFIG
# =========================================================
st.set_page_config(
    page_title="Athina Logistics Tool",
    page_icon="logo.png",
    layout="wide"
)

st.sidebar.image("logo.png", width=200)
st.sidebar.markdown("### Athina Logistics")
st.sidebar.caption("Global Access")

st.set_page_config(page_title="HS Code Checker CG", layout="wide")
st.title("HS Code Checker CG")
st.caption(
    "Upload one or more invoice files. "
    "The app checks HS codes in INVOICE!C20:C(SUM-1) and flags "
    "codes starting with 8714, or codes found in the watch list."
)

WATCHLIST = {
    "9405119090",
    "8414591500",
    "9505900000",
    "9503003990",
    "9503009990",
    "9405499090",
    "8714999089",
    "9405219090",
    "4819400000",
    "8305900000",
    "6217100090",
    "4821109000",
    "4202929190",
    "9506919000",
    "8510200000",
    "8504409590",
    "9505109000",
    "9620009900",
    "8544492000",
    "3926909790",
    "4820500000",
    "3924100090",
    "8714950000",
}

ALLOWED_EXTENSIONS = [".xlsx", ".xlsm", ".xltx", ".xltm"]


# =========================================================
# HELPERS
# =========================================================
def normalize_hs(value):
    if value is None:
        return ""

    s = str(value).strip()
    if not s:
        return ""

    s = s.replace("\n", " ").strip()
    s = re.sub(r"[^\d]", "", s)
    return s


def clean_text(value):
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def get_merged_value(ws, row, col):
    cell = ws.cell(row=row, column=col)

    if not isinstance(cell, MergedCell):
        return cell.value

    for merged_range in ws.merged_cells.ranges:
        if (
            merged_range.min_row <= row <= merged_range.max_row
            and merged_range.min_col <= col <= merged_range.max_col
        ):
            return ws.cell(merged_range.min_row, merged_range.min_col).value

    return None


def find_sum_row(ws, search_col=2):
    for r in range(1, ws.max_row + 1):
        val = get_merged_value(ws, r, search_col)
        if val is None:
            continue
        txt = str(val).strip().upper().replace(" ", "")
        if txt in {"SUM", "SUM:", "TOTAL", "TOTAL:"} or txt.startswith("SUM"):
            return r
    return None


def autofit_worksheet(ws):
    for col_cells in ws.columns:
        max_length = 0
        col_idx = col_cells[0].column
        col_letter = get_column_letter(col_idx)

        for cell in col_cells:
            try:
                value = "" if cell.value is None else str(cell.value)
                if len(value) > max_length:
                    max_length = len(value)
            except Exception:
                pass

        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 60)


def style_worksheet(ws):
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    ws.freeze_panes = "A2"

    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions

    autofit_worksheet(ws)


def build_excel_report(summary_df, issues_df):
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        issues_df.to_excel(writer, sheet_name="Issues", index=False)

        wb = writer.book
        ws_summary = wb["Summary"]
        ws_issues = wb["Issues"]

        style_worksheet(ws_summary)
        style_worksheet(ws_issues)

    output.seek(0)
    return output


def analyze_file(uploaded_file):
    issues = []

    filename = uploaded_file.name
    ext = Path(filename).suffix.lower()

    if ext not in ALLOWED_EXTENSIONS:
        return [], {
            "file": filename,
            "status": "ERROR",
            "message": f"Unsupported file type: {ext}",
            "checked_rows": 0,
            "issue_count": 0,
        }

    try:
        wb = load_workbook(filename=io.BytesIO(uploaded_file.getvalue()), data_only=True)
    except Exception as e:
        return [], {
            "file": filename,
            "status": "ERROR",
            "message": f"Cannot open file: {e}",
            "checked_rows": 0,
            "issue_count": 0,
        }

    if "INVOICE" not in wb.sheetnames:
        return [], {
            "file": filename,
            "status": "ERROR",
            "message": "Sheet 'INVOICE' not found",
            "checked_rows": 0,
            "issue_count": 0,
        }

    ws = wb["INVOICE"]
    sum_row = find_sum_row(ws, search_col=2)

    if not sum_row:
        return [], {
            "file": filename,
            "status": "ERROR",
            "message": "SUM row not found in column B",
            "checked_rows": 0,
            "issue_count": 0,
        }

    start_row = 20
    end_row = sum_row - 1

    if end_row < start_row:
        return [], {
            "file": filename,
            "status": "ERROR",
            "message": f"Invalid range: C{start_row}:C{end_row}",
            "checked_rows": 0,
            "issue_count": 0,
        }

    invoice_no = clean_text(get_merged_value(ws, 5, 3))
    checked_rows = 0

    for row in range(start_row, end_row + 1):
        raw_desc = get_merged_value(ws, row, 2)  # B
        raw_hs = get_merged_value(ws, row, 3)    # C

        description = clean_text(raw_desc)
        hs = normalize_hs(raw_hs)

        if not hs:
            continue

        checked_rows += 1
        reasons = []

        if hs[:2] == "8714":
            reasons.append("HS starts with 8714")

        if hs in WATCHLIST:
            reasons.append("HS found in watch list")

        if reasons:
            issues.append(
                {
                    "File": filename,
                    "Invoice No": invoice_no,
                    "Sheet": "INVOICE",
                    "Row": row,
                    "Cell": f"C{row}",
                    "Description": description,
                    "HS Code": hs,
                    "Reason": " | ".join(reasons),
                    "Location": f"{filename} | Invoice {invoice_no} | INVOICE | C{row} | {description}",
                }
            )

    status = "OK" if not issues else "FLAGGED"
    message = "No issue found" if not issues else f"{len(issues)} issue(s) found"

    return issues, {
        "file": filename,
        "invoice_no": invoice_no,
        "status": status,
        "message": message,
        "checked_rows": checked_rows,
        "issue_count": len(issues),
    }


# =========================================================
# UI
# =========================================================
st.subheader("Upload invoice files")

uploaded_files = st.file_uploader(
    "Select Excel invoice files",
    type=["xlsx", "xlsm", "xltx", "xltm"],
    accept_multiple_files=True,
)

col1, col2 = st.columns([1, 1])
with col1:
    run_btn = st.button("Run HS check", type="primary")
with col2:
    show_watchlist = st.checkbox("Show watch list", value=False)

if show_watchlist:
    st.markdown("### Watch list")
    watch_df = pd.DataFrame({"HS Code": sorted(WATCHLIST)})
    st.dataframe(watch_df, use_container_width=True, hide_index=True)

if run_btn:
    if not uploaded_files:
        st.warning("Please upload at least one file.")
        st.stop()

    all_issues = []
    all_summaries = []

    progress = st.progress(0)
    status_box = st.empty()

    total_files = len(uploaded_files)

    for idx, f in enumerate(uploaded_files, start=1):
        status_box.info(f"Checking {idx}/{total_files}: {f.name}")
        issues, summary = analyze_file(f)
        all_issues.extend(issues)
        all_summaries.append(summary)
        progress.progress(idx / total_files)

    status_box.success("Check completed.")

    summary_df = pd.DataFrame(all_summaries)
    issues_df = pd.DataFrame(all_issues)

    if summary_df.empty:
        summary_df = pd.DataFrame(
            columns=["file", "invoice_no", "status", "message", "checked_rows", "issue_count"]
        )

    if issues_df.empty:
        issues_df = pd.DataFrame(
            columns=[
                "File",
                "Invoice No",
                "Sheet",
                "Row",
                "Cell",
                "Description",
                "HS Code",
                "Reason",
                "Location",
            ]
        )

    st.markdown("## Summary")
    st.dataframe(summary_df, use_container_width=True, hide_index=True)

    total_flagged_files = int((summary_df["status"] == "FLAGGED").sum()) if "status" in summary_df.columns else 0
    total_error_files = int((summary_df["status"] == "ERROR").sum()) if "status" in summary_df.columns else 0
    total_issues = len(issues_df)

    c1, c2, c3 = st.columns(3)
    c1.metric("Files checked", len(summary_df))
    c2.metric("Flagged files", total_flagged_files)
    c3.metric("Total issues", total_issues)

    if total_error_files > 0:
        st.error(f"{total_error_files} file(s) could not be checked.")

    st.markdown("## Issues found")
    if issues_df.empty:
        st.success("No HS issue found.")
    else:
        st.dataframe(issues_df, use_container_width=True, hide_index=True)

    report_file = build_excel_report(summary_df, issues_df)

    st.download_button(
        label="Download Excel report",
        data=report_file,
        file_name="hs_code_check_report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
